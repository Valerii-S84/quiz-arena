"""IO helpers for the quiz bank ambiguity scan."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore


def read_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        columns = list(reader.fieldnames or [])
        rows = []
        for row_idx, row in enumerate(reader, start=2):
            row["_row"] = row_idx
            rows.append(row)
        return columns, rows


def read_xlsx(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    columns = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {"_row": r}
        empty = True
        for c, col in enumerate(columns, start=1):
            value = ws.cell(r, c).value
            row[col] = value
            if value not in (None, ""):
                empty = False
        if not empty:
            rows.append(row)
    return columns, rows


def read_table(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if path.suffix.lower() == ".csv":
        return read_csv(path)
    return read_xlsx(path)

#!/usr/bin/env python3
"""Run structural and quality audits for quiz bank tables."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None


REQUIRED_COLUMNS = [
    "quiz_id",
    "question",
    "option_1",
    "option_2",
    "option_3",
    "option_4",
    "correct_option_id",
    "correct_answer",
    "explanation",
    "key",
    "level",
    "category",
]

OPTION_COLUMNS = ["option_1", "option_2", "option_3", "option_4"]
TEXT_CHECK_COLUMNS = [
    "question",
    "option_1",
    "option_2",
    "option_3",
    "option_4",
    "correct_answer",
    "explanation",
    "key",
]
DATE_COLUMNS = ["created_at", "published_at"]


@dataclass
class TableData:
    columns: list[str]
    rows: list[dict[str, Any]]
    parser: str
    warnings: list[str]


def normalize(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("‑", "-").replace("–", "-").replace("—", "-")
    return text


def parse_date(value: Any) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    if not text:
        return True
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        datetime.fromisoformat(text)
        return True
    except ValueError:
        return False


def read_csv_table(path: Path) -> TableData:
    warnings: list[str] = []
    last_error: Exception | None = None
    encodings = ["utf-8-sig", "utf-8", "cp1251"]
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                if reader.fieldnames is None:
                    return TableData([], [], "csv", [f"{path.name}: missing header"])
                columns = [name.strip() if name else "" for name in reader.fieldnames]
                rows: list[dict[str, Any]] = []
                for index, row in enumerate(reader, start=2):
                    row_data = {col: row.get(col) for col in columns}
                    row_data["_row"] = index
                    rows.append(row_data)
                return TableData(columns, rows, "csv", warnings)
        except Exception as err:
            last_error = err
    message = f"{path.name}: failed to read CSV ({last_error})"
    return TableData([], [], "csv", [message])


def read_xlsx_table(path: Path) -> TableData:
    warnings: list[str] = []
    if load_workbook is None:
        return TableData([], [], "xlsx", [f"{path.name}: openpyxl is not installed"])
    workbook = load_workbook(path, data_only=True)
    if not workbook.sheetnames:
        return TableData([], [], "xlsx", [f"{path.name}: workbook has no sheets"])
    sheet = workbook[workbook.sheetnames[0]]
    header = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    columns = [str(value).strip() if value is not None else "" for value in header]
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, sheet.max_row + 1):
        row_data: dict[str, Any] = {"_row": row_idx}
        is_empty = True
        for col_idx, col_name in enumerate(columns, start=1):
            value = sheet.cell(row_idx, col_idx).value
            row_data[col_name] = value
            if value not in (None, ""):
                is_empty = False
        if not is_empty:
            rows.append(row_data)
    return TableData(columns, rows, "xlsx", warnings)


def read_table(path: Path) -> TableData:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_table(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_table(path)
    return TableData([], [], "unknown", [f"{path.name}: unsupported extension {suffix}"])


def pick(row: dict[str, Any], column: str) -> Any:
    return row.get(column)


def audit_table(path: Path, data: TableData) -> dict[str, Any]:
    columns = data.columns
    rows = data.rows
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in columns]

    missing_required_counts: Counter[str] = Counter()
    trailing_space_samples: list[dict[str, Any]] = []
    double_space_samples: list[dict[str, Any]] = []
    question_punct_issues: list[int] = []
    explanation_punct_issues: list[int] = []
    date_parse_issues: list[dict[str, Any]] = []
    status_distribution: Counter[str] = Counter()

    bad_correct_option_id: list[dict[str, Any]] = []
    mismatch_correct_answer: list[dict[str, Any]] = []
    answer_not_in_options: list[dict[str, Any]] = []

    question_groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    key_groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in rows:
        row_id = int(row["_row"])

        for col in REQUIRED_COLUMNS:
            if col not in columns:
                continue
            value = pick(row, col)
            if value is None or (isinstance(value, str) and value.strip() == ""):
                missing_required_counts[col] += 1

        for col in TEXT_CHECK_COLUMNS:
            if col not in columns:
                continue
            value = pick(row, col)
            if not isinstance(value, str):
                continue
            if value != value.strip() and len(trailing_space_samples) < 50:
                trailing_space_samples.append(
                    {"row": row_id, "column": col, "value": value}
                )
            if "  " in value and len(double_space_samples) < 50:
                double_space_samples.append({"row": row_id, "column": col})

        question = pick(row, "question")
        if isinstance(question, str):
            trimmed = question.strip()
            if trimmed and not trimmed.endswith("?"):
                question_punct_issues.append(row_id)

        explanation = pick(row, "explanation")
        if isinstance(explanation, str):
            trimmed = explanation.strip()
            if trimmed and not re.search(r"[.!?]$", trimmed):
                explanation_punct_issues.append(row_id)

        for date_col in DATE_COLUMNS:
            if date_col in columns and not parse_date(pick(row, date_col)):
                if len(date_parse_issues) < 50:
                    date_parse_issues.append(
                        {"row": row_id, "column": date_col, "value": pick(row, date_col)}
                    )

        if "status" in columns:
            status_distribution[normalize(pick(row, "status")) or "(empty)"] += 1

        if "question" in columns:
            question_groups[normalize(pick(row, "question"))].append(row)
        if "key" in columns:
            key_groups[normalize(pick(row, "key"))].append(row)

        if "correct_option_id" in columns:
            raw_id = pick(row, "correct_option_id")
            valid = False
            try:
                parsed = int(str(raw_id).strip())
                if str(raw_id).strip() in {"0", "1", "2", "3"} or isinstance(raw_id, int):
                    valid = 0 <= parsed <= 3
            except Exception:
                valid = False
                parsed = -1
            if not valid:
                if len(bad_correct_option_id) < 50:
                    bad_correct_option_id.append({"row": row_id, "value": raw_id})
                continue

            options = [pick(row, option_col) for option_col in OPTION_COLUMNS]
            correct_answer = pick(row, "correct_answer")
            selected = options[parsed] if parsed < len(options) else None
            if normalize(correct_answer) != normalize(selected):
                if len(mismatch_correct_answer) < 50:
                    mismatch_correct_answer.append(
                        {
                            "row": row_id,
                            "correct_option_id": parsed,
                            "correct_answer": correct_answer,
                            "selected_option": selected,
                        }
                    )
            if normalize(correct_answer) not in [normalize(option) for option in options]:
                if len(answer_not_in_options) < 50:
                    answer_not_in_options.append(
                        {"row": row_id, "correct_answer": correct_answer}
                    )

    duplicate_question_groups: list[dict[str, Any]] = []
    ambiguous_question_groups: list[dict[str, Any]] = []
    for question, group_rows in question_groups.items():
        if not question or len(group_rows) < 2:
            continue
        row_ids = [int(row["_row"]) for row in group_rows]
        answers = {normalize(pick(row, "correct_answer")) for row in group_rows}
        duplicate_question_groups.append(
            {
                "question": pick(group_rows[0], "question"),
                "rows": row_ids,
                "answers": sorted(answers),
            }
        )
        if len(answers) > 1:
            ambiguous_question_groups.append(
                {
                    "question": pick(group_rows[0], "question"),
                    "rows": row_ids,
                    "answers": sorted(answers),
                }
            )

    duplicate_key_groups = [
        {"key": pick(group_rows[0], "key"), "rows": [int(row["_row"]) for row in group_rows]}
        for key, group_rows in key_groups.items()
        if key and len(group_rows) > 1
    ]

    non_ready_count = sum(
        count
        for status, count in status_distribution.items()
        if status not in {"ready", "(empty)"}
    )

    critical_count = (
        len(missing_columns)
        + sum(missing_required_counts.values())
        + len(bad_correct_option_id)
        + len(mismatch_correct_answer)
        + len(answer_not_in_options)
    )
    high_count = len(ambiguous_question_groups)
    medium_count = (
        len(duplicate_question_groups)
        + len(duplicate_key_groups)
        + len(trailing_space_samples)
        + len(double_space_samples)
        + len(date_parse_issues)
        + non_ready_count
    )

    if critical_count == 0 and high_count == 0 and medium_count == 0:
        readiness = "ready"
    elif critical_count == 0 and high_count == 0:
        readiness = "ready_with_cleanup"
    else:
        readiness = "needs_fix"

    return {
        "file": str(path),
        "parser": data.parser,
        "warnings": data.warnings,
        "row_count": len(rows),
        "column_count": len(columns),
        "columns": columns,
        "missing_required_columns": missing_columns,
        "missing_required_counts": dict(missing_required_counts),
        "invalid_correct_option_id_count": len(bad_correct_option_id),
        "invalid_correct_option_id_samples": bad_correct_option_id,
        "mismatch_correct_answer_count": len(mismatch_correct_answer),
        "mismatch_correct_answer_samples": mismatch_correct_answer,
        "answer_not_in_options_count": len(answer_not_in_options),
        "answer_not_in_options_samples": answer_not_in_options,
        "duplicate_question_group_count": len(duplicate_question_groups),
        "duplicate_question_groups": duplicate_question_groups,
        "ambiguous_question_group_count": len(ambiguous_question_groups),
        "ambiguous_question_groups": ambiguous_question_groups,
        "duplicate_key_group_count": len(duplicate_key_groups),
        "duplicate_key_groups": duplicate_key_groups,
        "trailing_space_issue_count": len(trailing_space_samples),
        "trailing_space_samples": trailing_space_samples,
        "double_space_issue_count": len(double_space_samples),
        "double_space_samples": double_space_samples,
        "question_terminal_punctuation_issue_count": len(question_punct_issues),
        "question_terminal_punctuation_issue_rows": question_punct_issues[:100],
        "explanation_terminal_punctuation_issue_count": len(explanation_punct_issues),
        "explanation_terminal_punctuation_issue_rows": explanation_punct_issues[:100],
        "date_parse_issue_count": len(date_parse_issues),
        "date_parse_issue_samples": date_parse_issues,
        "status_distribution": dict(status_distribution),
        "readiness": readiness,
        "severity": {
            "critical": critical_count,
            "high": high_count,
            "medium": medium_count,
        },
    }


def make_action_plan(file_report: dict[str, Any]) -> list[str]:
    actions: list[str] = []
    if file_report["missing_required_columns"]:
        actions.append("Add missing required columns to match ingest schema.")
    if file_report["missing_required_counts"]:
        actions.append("Fill empty required cells in existing rows.")
    if file_report["invalid_correct_option_id_count"] > 0:
        actions.append("Fix invalid correct_option_id values (must be 0..3).")
    if file_report["mismatch_correct_answer_count"] > 0:
        actions.append("Align correct_answer with option indexed by correct_option_id.")
    if file_report["answer_not_in_options_count"] > 0:
        actions.append("Ensure correct_answer exists among option_1..option_4.")
    if file_report["ambiguous_question_group_count"] > 0:
        actions.append("Rewrite duplicated stems with different correct answers.")
    if file_report["duplicate_question_group_count"] > 0:
        actions.append("Reduce exact duplicate questions or mark intended repetitions.")
    if file_report["duplicate_key_group_count"] > 0:
        actions.append("Make key values unique for stable indexing.")
    if file_report["trailing_space_issue_count"] > 0 or file_report["double_space_issue_count"] > 0:
        actions.append("Normalize whitespace in questions, options, explanations and keys.")
    if file_report["date_parse_issue_count"] > 0:
        actions.append("Normalize timestamp format to ISO 8601.")
    if file_report["question_terminal_punctuation_issue_count"] > 0:
        actions.append("Review question punctuation and end marks.")
    if file_report["explanation_terminal_punctuation_issue_count"] > 0:
        actions.append("Ensure explanations end with punctuation.")
    if not actions:
        actions.append("No blocking issues detected. Keep file under regression QA.")
    return actions


def build_markdown(report: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append("# QuizBank Audit Report")
    lines.append("")
    lines.append(f"- Generated at: `{report['generated_at']}`")
    lines.append(f"- Files audited: `{report['summary']['file_count']}`")
    lines.append(f"- Total rows: `{report['summary']['row_count']}`")
    lines.append("")
    lines.append("## Checklist by file")
    lines.append("")
    lines.append(
        "| File | Rows | Readiness | Critical | High | Medium | Ambiguous groups | Duplicate question groups |"
    )
    lines.append("|---|---:|---|---:|---:|---:|---:|---:|")
    for file_report in report["files"]:
        lines.append(
            "| "
            f"{Path(file_report['file']).name} | "
            f"{file_report['row_count']} | "
            f"{file_report['readiness']} | "
            f"{file_report['severity']['critical']} | "
            f"{file_report['severity']['high']} | "
            f"{file_report['severity']['medium']} | "
            f"{file_report['ambiguous_question_group_count']} | "
            f"{file_report['duplicate_question_group_count']} |"
        )

    lines.append("")
    lines.append("## Per-file action plan")
    lines.append("")
    for file_report in report["files"]:
        lines.append(f"### {Path(file_report['file']).name}")
        lines.append("")
        lines.append(
            "- Status: "
            f"`{file_report['readiness']}` "
            f"(critical `{file_report['severity']['critical']}`, "
            f"high `{file_report['severity']['high']}`, "
            f"medium `{file_report['severity']['medium']}`)"
        )
        actions = make_action_plan(file_report)
        for action in actions:
            lines.append(f"- {action}")
        lines.append("")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit quiz bank files.")
    parser.add_argument("--input-dir", default="QuizBank", help="Directory with quiz files.")
    parser.add_argument(
        "--output-json",
        default="reports/quizbank_audit_report.json",
        help="Path to JSON report file.",
    )
    parser.add_argument(
        "--output-md",
        default="reports/quizbank_audit_report.md",
        help="Path to Markdown report file.",
    )
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    files = sorted(
        path
        for path in input_dir.iterdir()
        if path.is_file() and path.suffix.lower() in {".csv", ".xlsx", ".xlsm"}
    )

    file_reports = []
    for file_path in files:
        table = read_table(file_path)
        file_reports.append(audit_table(file_path, table))

    summary = {
        "file_count": len(file_reports),
        "row_count": sum(item["row_count"] for item in file_reports),
        "ready_count": sum(1 for item in file_reports if item["readiness"] == "ready"),
        "ready_with_cleanup_count": sum(
            1 for item in file_reports if item["readiness"] == "ready_with_cleanup"
        ),
        "needs_fix_count": sum(1 for item in file_reports if item["readiness"] == "needs_fix"),
        "critical_total": sum(item["severity"]["critical"] for item in file_reports),
        "high_total": sum(item["severity"]["high"] for item in file_reports),
        "medium_total": sum(item["severity"]["medium"] for item in file_reports),
    }
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "input_dir": str(input_dir),
        "summary": summary,
        "files": file_reports,
    }

    output_json = Path(args.output_json)
    output_md = Path(args.output_md)
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_md.parent.mkdir(parents=True, exist_ok=True)

    output_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    output_md.write_text(build_markdown(report), encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"JSON report: {output_json}")
    print(f"Markdown report: {output_md}")


if __name__ == "__main__":
    main()

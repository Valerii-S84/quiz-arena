#!/usr/bin/env python3
"""Focused scan for duplicates/repeats and potential multi-correct ambiguity."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore
from rapidfuzz import fuzz


OPTION_COLUMNS = ["option_1", "option_2", "option_3", "option_4"]
CUE_PREFIXES = [
    "was passt logisch?",
    "was passt inhaltlich?",
    "was passt semantisch?",
    "sinnvoll ergänzen:",
    "sinnvoll ergaenzen:",
    "inhaltlich richtig:",
    "inhaltlich logisch:",
    "inhaltlich stimmig ergänzen:",
    "inhaltlich stimmig ergaenzen:",
    "logisch ergänzen:",
    "logisch ergaenzen:",
]


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("‑", "-").replace("–", "-").replace("—", "-")
    text = re.sub(r"\s+", " ", text)
    return text


def question_signature(question: str) -> str:
    text = norm(question)
    changed = True
    while changed:
        changed = False
        for cue in CUE_PREFIXES:
            if text.startswith(cue + " "):
                text = text[len(cue) + 1 :].strip()
                changed = True
            elif text.startswith(cue):
                text = text[len(cue) :].strip()
                changed = True
        text = text.lstrip(":-,; ").strip()
    text = text.replace("_____", "")
    text = text.replace("___", "")
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"[\"“”„']", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def key_family(value: Any) -> str:
    text = norm(value)
    prev = None
    while prev != text:
        prev = text
        text = re.sub(r"(?:_|-)v\d+$", "", text, flags=re.I)
        text = re.sub(r"(?:_|-)\d+$", "", text)
    return text


def read_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        columns = list(reader.fieldnames or [])
        rows = []
        for row_idx, row in enumerate(reader, start=2):
            row["_row"] = row_idx
            rows.append(row)
        return columns, rows


def read_xlsx(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    columns = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {"_row": r}
        empty = True
        for c, col in enumerate(columns, start=1):
            value = ws.cell(r, c).value
            row[col] = value
            if value not in (None, ""):
                empty = False
        if not empty:
            rows.append(row)
    return columns, rows


def read_table(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if path.suffix.lower() == ".csv":
        return read_csv(path)
    return read_xlsx(path)


def scan_file(path: Path) -> dict[str, Any]:
    columns, rows = read_table(path)
    if "question" not in columns:
        return {
            "file": str(path),
            "row_count": len(rows),
            "exact_duplicate_groups": [],
            "exact_conflict_groups": [],
            "signature_conflict_groups": [],
            "fuzzy_conflicts": [],
            "same_row_multi_logical_candidates": [],
        }

    question_groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    signature_groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    key_family_groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    repeats_same_answer: list[dict[str, Any]] = []
    exact_conflicts: list[dict[str, Any]] = []
    signature_conflicts: list[dict[str, Any]] = []
    fuzzy_conflicts: list[dict[str, Any]] = []
    same_row_candidates: list[dict[str, Any]] = []

    for row in rows:
        q = norm(row.get("question"))
        if q:
            question_groups[q].append(row)
            signature_groups[question_signature(str(row.get("question") or ""))].append(row)
        if "key" in columns:
            family = key_family(row.get("key"))
            if family:
                key_family_groups[family].append(row)

        options = [norm(row.get(col)) for col in OPTION_COLUMNS if col in columns]
        if len(options) >= 2 and len(set(options)) < len(options):
            same_row_candidates.append(
                {
                    "row": int(row["_row"]),
                    "reason": "duplicate options after normalization",
                    "question": row.get("question"),
                    "options": [row.get(col) for col in OPTION_COLUMNS if col in columns],
                }
            )

    for q, group in question_groups.items():
        if len(group) < 2:
            continue
        answers = sorted({norm(g.get("correct_answer")) for g in group})
        item = {
            "question": group[0].get("question"),
            "rows": [int(g["_row"]) for g in group],
            "answers": answers,
        }
        if len(answers) == 1:
            repeats_same_answer.append(item)
        else:
            exact_conflicts.append(item)

    for sig, group in signature_groups.items():
        if len(group) < 2 or not sig:
            continue
        answers = sorted({norm(g.get("correct_answer")) for g in group})
        if len(answers) > 1:
            signature_conflicts.append(
                {
                    "signature": sig,
                    "rows": [int(g["_row"]) for g in group],
                    "questions": [g.get("question") for g in group],
                    "answers": answers,
                }
            )

    template_repeat_groups: list[dict[str, Any]] = []
    for family, group in key_family_groups.items():
        if len(group) < 2:
            continue
        answers = sorted({norm(g.get("correct_answer")) for g in group})
        template_repeat_groups.append(
            {
                "family": family,
                "size": len(group),
                "rows": [int(g["_row"]) for g in group],
                "questions": [g.get("question") for g in group][:10],
                "answers": answers,
                "answer_count": len(answers),
            }
        )
    template_repeat_groups = sorted(template_repeat_groups, key=lambda x: x["size"], reverse=True)

    q_items = [(int(row["_row"]), norm(row.get("question")), norm(row.get("correct_answer"))) for row in rows]
    for i, (r1, q1, a1) in enumerate(q_items):
        if not q1:
            continue
        for r2, q2, a2 in q_items[i + 1 :]:
            if not q2 or a1 == a2:
                continue
            score = fuzz.ratio(q1, q2)
            if score >= 96:
                fuzzy_conflicts.append(
                    {
                        "row_1": r1,
                        "row_2": r2,
                        "score": score,
                        "question_1": q1,
                        "question_2": q2,
                        "answer_1": a1,
                        "answer_2": a2,
                    }
                )
    fuzzy_conflicts = sorted(fuzzy_conflicts, key=lambda x: x["score"], reverse=True)[:100]

    return {
        "file": str(path),
        "row_count": len(rows),
        "exact_duplicate_groups": repeats_same_answer,
        "exact_conflict_groups": exact_conflicts,
        "signature_conflict_groups": signature_conflicts[:100],
        "fuzzy_conflicts": fuzzy_conflicts,
        "same_row_multi_logical_candidates": same_row_candidates[:100],
        "template_repeat_groups": template_repeat_groups,
    }


def build_md(report: dict[str, Any]) -> str:
    lines = ["# QuizBank Duplicate & Ambiguity Scan", ""]
    lines.append(f"- Generated at: `{report['generated_at']}`")
    lines.append(f"- Files scanned: `{len(report['files'])}`")
    lines.append("")
    lines.append("| File | Rows | Exact duplicate groups | Exact conflict groups | Signature conflict groups | Fuzzy conflicts | Same-row candidates | Template repeat groups |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|---:|")
    for item in report["files"]:
        lines.append(
            f"| {Path(item['file']).name} | {item['row_count']} | "
            f"{len(item['exact_duplicate_groups'])} | {len(item['exact_conflict_groups'])} | "
            f"{len(item['signature_conflict_groups'])} | {len(item['fuzzy_conflicts'])} | "
            f"{len(item['same_row_multi_logical_candidates'])} | "
            f"{len(item['template_repeat_groups'])} |"
        )
    lines.append("")
    lines.append("## Conflict Details")
    lines.append("")
    for item in report["files"]:
        if (
            not item["exact_conflict_groups"]
            and not item["signature_conflict_groups"]
            and not item["fuzzy_conflicts"]
            and not item["same_row_multi_logical_candidates"]
            and not item["template_repeat_groups"]
        ):
            continue
        lines.append(f"### {Path(item['file']).name}")
        lines.append("")
        for grp in item["exact_conflict_groups"]:
            lines.append(
                f"- Exact conflict rows `{grp['rows']}` | answers `{grp['answers']}` | question: `{grp['question']}`"
            )
        for grp in item["signature_conflict_groups"][:10]:
            lines.append(
                f"- Signature conflict rows `{grp['rows']}` | answers `{grp['answers']}` | signature: `{grp['signature']}`"
            )
        for pair in item["fuzzy_conflicts"][:10]:
            lines.append(
                f"- Fuzzy conflict `{pair['row_1']}/{pair['row_2']}` score `{pair['score']}` "
                f"| answers `{pair['answer_1']}/{pair['answer_2']}`"
            )
        for candidate in item["same_row_multi_logical_candidates"][:10]:
            lines.append(
                f"- Same-row candidate row `{candidate['row']}` | reason: `{candidate['reason']}`"
            )
        for group in item["template_repeat_groups"][:10]:
            lines.append(
                f"- Template-repeat family `{group['family']}` | size `{group['size']}` "
                f"| rows `{group['rows']}` | answer_count `{group['answer_count']}`"
            )
        lines.append("")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Scan quiz bank for duplicates and ambiguity.")
    parser.add_argument("--input-dir", default="QuizBank")
    parser.add_argument("--output-json", default="reports/quizbank_ambiguity_scan.json")
    parser.add_argument("--output-md", default="reports/quizbank_ambiguity_scan.md")
    args = parser.parse_args()

    files = sorted(
        [
            p
            for p in Path(args.input_dir).iterdir()
            if p.is_file() and p.suffix.lower() in {".csv", ".xlsx", ".xlsm"}
        ]
    )
    results = [scan_file(p) for p in files]
    summary = {
        "file_count": len(results),
        "exact_duplicate_groups_total": sum(len(r["exact_duplicate_groups"]) for r in results),
        "exact_conflict_groups_total": sum(len(r["exact_conflict_groups"]) for r in results),
        "signature_conflict_groups_total": sum(len(r["signature_conflict_groups"]) for r in results),
        "fuzzy_conflicts_total": sum(len(r["fuzzy_conflicts"]) for r in results),
        "same_row_candidates_total": sum(
            len(r["same_row_multi_logical_candidates"]) for r in results
        ),
        "template_repeat_groups_total": sum(len(r["template_repeat_groups"]) for r in results),
        "template_repeat_rows_total": sum(
            sum(group["size"] for group in r["template_repeat_groups"]) for r in results
        ),
    }
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": summary,
        "files": results,
    }

    out_json = Path(args.output_json)
    out_md = Path(args.output_md)
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_md.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    out_md.write_text(build_md(report), encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"JSON: {out_json}")
    print(f"MD: {out_md}")


if __name__ == "__main__":
    main()

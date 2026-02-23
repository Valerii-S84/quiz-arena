from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None


@dataclass
class TableData:
    columns: list[str]
    rows: list[dict[str, Any]]
    parser: str
    warnings: list[str]


def normalize(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("‑", "-").replace("–", "-").replace("—", "-")
    return text


def parse_date(value: Any) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    if not text:
        return True
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        datetime.fromisoformat(text)
        return True
    except ValueError:
        return False


def read_csv_table(path: Path) -> TableData:
    warnings: list[str] = []
    last_error: Exception | None = None
    encodings = ["utf-8-sig", "utf-8", "cp1251"]
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                if reader.fieldnames is None:
                    return TableData([], [], "csv", [f"{path.name}: missing header"])
                columns = [name.strip() if name else "" for name in reader.fieldnames]
                rows: list[dict[str, Any]] = []
                for index, row in enumerate(reader, start=2):
                    row_data = {col: row.get(col) for col in columns}
                    row_data["_row"] = index
                    rows.append(row_data)
                return TableData(columns, rows, "csv", warnings)
        except Exception as err:
            last_error = err
    message = f"{path.name}: failed to read CSV ({last_error})"
    return TableData([], [], "csv", [message])


def read_xlsx_table(path: Path) -> TableData:
    warnings: list[str] = []
    if load_workbook is None:
        return TableData([], [], "xlsx", [f"{path.name}: openpyxl is not installed"])
    workbook = load_workbook(path, data_only=True)
    if not workbook.sheetnames:
        return TableData([], [], "xlsx", [f"{path.name}: workbook has no sheets"])
    sheet = workbook[workbook.sheetnames[0]]
    header = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    columns = [str(value).strip() if value is not None else "" for value in header]
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, sheet.max_row + 1):
        row_data: dict[str, Any] = {"_row": row_idx}
        is_empty = True
        for col_idx, col_name in enumerate(columns, start=1):
            value = sheet.cell(row_idx, col_idx).value
            row_data[col_name] = value
            if value not in (None, ""):
                is_empty = False
        if not is_empty:
            rows.append(row_data)
    return TableData(columns, rows, "xlsx", warnings)


def read_table(path: Path) -> TableData:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_table(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_table(path)
    return TableData([], [], "unknown", [f"{path.name}: unsupported extension {suffix}"])
